from collections import Counter, defaultdict
import openpyxl
import os

# Load the workbook and select the worksheet
workbook = openpyxl.load_workbook('PATH/TO/YOUR/EXCELFILE.xlsx')
worksheet = workbook['Combined_all_tags_UPDATE']

# Define the specific value to filter rows by in column B
specific_value = 'EXCEL_FILE_NAME.xlsx'

# Filter rows by the specific value in column B
filtered_rows = [row for row in worksheet.iter_rows(min_row=2, values_only=True) if row[1] == specific_value]

# Initialize a Counter to count occurrences of each value
value_counter = Counter()
# Initialize a defaultdict to keep track of strings in column D
value_strings = defaultdict(list)

# Iterate over the filtered rows
for row in filtered_rows:
    # Get the string from column D
    col_d_string = row[3]
    # Iterate over the columns starting from column E
    for value in row[4:]:
        if value is not None:
            value_counter[value] += 1
            value_strings[value].append(col_d_string)

# Set the label to the specific value
label = specific_value

# Define the path for the results workbook
results_workbook_path = 'PATH/TO/YOUR/Results_Workbook.xlsx'

# Check if the results workbook exists, if not create a new one
if os.path.exists(results_workbook_path):
    results_workbook = openpyxl.load_workbook(results_workbook_path)
else:
    results_workbook = openpyxl.Workbook()
    # Remove the default sheet created by openpyxl
    results_workbook.remove(results_workbook.active)

# Create a new worksheet with the label as the title followed by ' VG'
new_worksheet = results_workbook.create_sheet(title=f"{label} VG")

# Write the label to cell A1
new_worksheet['A1'] = f"Results labelled with: {label}"

# Write the header to the new worksheet
new_worksheet.append(["Value", "Count", "Strings in Column D", "Definition"])

# Load the SEMTAGS worksheet
semtags_worksheet = results_workbook['SEMTAGS']

# Create a dictionary to store SEMTAGS definitions
semtags_dict = {row[0]: row[1] for row in semtags_worksheet.iter_rows(min_row=2, values_only=True)}

# Write the results to the new worksheet
for value, count in value_counter.items():
    strings = ', '.join(value_strings[value])
    definition = semtags_dict.get(value, "No definition found")
    new_worksheet.append([value, count, strings, definition])

# Save the results workbook
results_workbook.save(results_workbook_path)