import os
import openpyxl
from collections import defaultdict
import csv

def process_workbook(workbook_path):
    print(f"Processing workbook: {workbook_path}")
    workbook = openpyxl.load_workbook(workbook_path)
    
    results = {}
    
    # Process 'extra' worksheet if it exists
    extra_metaphor_count = 0
    if 'extra' in workbook.sheetnames:
        extra_sheet = workbook['extra']
        for row in extra_sheet.iter_rows(min_row=2):  # Skip header row
            if row[4].value and str(row[4].value).upper() in ('Y', 'O'):  # Check column E (index 4)
                extra_metaphor_count += 1
        print(f"Found {extra_metaphor_count} metaphors in 'extra' worksheet")
    
    results['extra_metaphors'] = extra_metaphor_count
    
    # Find regular worksheets and their corresponding sample worksheets
    regular_sheets = [name for name in workbook.sheetnames 
                     if not name.endswith('_20%') 
                     and name != 'Coding List' 
                     and name != 'Coding Lists'  # Exclude both naming variants
                     and name != 'Coding_lists'  # Exclude both naming variants
                     and name != 'extra'
                     and not name.endswith('_lf')]
    
    for sheet_name in regular_sheets:
        # Skip any sheets with 'Coding List' in the name (additional check)
        if 'coding list' in sheet_name.lower():
            continue
            
        print(f"Processing regular worksheet: {sheet_name}")
        sheet = workbook[sheet_name]
        
        # Count total rows in regular worksheet (excluding header)
        row_count = sheet.max_row - 1  # Subtract header row
        results[sheet_name] = {'total_rows': row_count}
        
        # Count occurrences of each string in column C
        string_counts = defaultdict(int)
        for row in sheet.iter_rows(min_row=2):  # Skip header
            if row[2].value:  # Column C (index 2)
                string_value = str(row[2].value)
                string_counts[string_value] += 1
        
        # Find low frequency strings (≤ 20 occurrences)
        low_freq_strings = [string for string, count in string_counts.items() if count <= 20]
        results[sheet_name]['low_freq_strings'] = low_freq_strings
        print(f"Found {len(low_freq_strings)} low frequency strings in {sheet_name}")
        
        # Look for corresponding sample worksheet
        sample_sheet_name = f"{sheet_name}_20%"
        if sample_sheet_name in workbook.sheetnames:
            print(f"Processing sample worksheet: {sample_sheet_name}")
            sample_sheet = workbook[sample_sheet_name]
            
            # Create a new low frequency worksheet
            lf_sheet_name = f"{sheet_name}_lf"
            if lf_sheet_name in workbook.sheetnames:
                # Remove existing sheet if it exists
                del workbook[lf_sheet_name]
            
            lf_sheet = workbook.create_sheet(lf_sheet_name)
            
            # Copy header row
            header = [cell.value for cell in sample_sheet[1]]
            lf_sheet.append(header)
            
            # Find rows with low frequency strings and move them to the new sheet
            rows_to_remove = []
            lf_rows_count = 0  # Count of LF rows
            for i, row in enumerate(sample_sheet.iter_rows(min_row=2), start=2):
                if row[2].value and str(row[2].value) in low_freq_strings:
                    # Copy row to low frequency sheet
                    lf_sheet.append([cell.value for cell in row])
                    rows_to_remove.append(i)
                    lf_rows_count += 1
            
            # Store the count of LF rows
            results[sheet_name]['total_lf_rows'] = lf_rows_count
            
            # Calculate and store HF rows (total rows minus LF rows)
            results[sheet_name]['total_hf_rows'] = row_count - lf_rows_count
            
            print(f"Total LF rows: {lf_rows_count}")
            print(f"Total HF rows: {row_count - lf_rows_count}")
            
            # Count metaphors in low frequency sheet
            lf_metaphor_count = 0
            for row in lf_sheet.iter_rows(min_row=2):  # Skip header
                if row[4].value and str(row[4].value).upper() in ('Y', 'O'):  # Column E (index 4)
                    lf_metaphor_count += 1
            
            results[sheet_name]['low_freq_metaphors'] = lf_metaphor_count
            print(f"Found {lf_metaphor_count} metaphors in low frequency rows")
            
            # Remove rows from the sample sheet (in reverse order to maintain indices)
            for row_idx in sorted(rows_to_remove, reverse=True):
                sample_sheet.delete_rows(row_idx)
            
            # Count rows and metaphors in sample sheet AFTER removing low frequency types
            sample_remaining_rows = sample_sheet.max_row - 1  # Excluding header
            
            # Count metaphors in remaining sample rows
            sample_metaphor_count = 0
            for row in sample_sheet.iter_rows(min_row=2):  # Skip header
                if row[4].value and str(row[4].value).upper() in ('Y', 'O'):  # Column E (index 4)
                    sample_metaphor_count += 1
            
            results[sheet_name]['sample_remaining_rows'] = sample_remaining_rows
            results[sheet_name]['sample_metaphors'] = sample_metaphor_count
            # --- Added: scale-up estimate for high-frequency (HF) sample + final estimate (excluding 'extra') ---
            # Scale up based on the proportion of sample rows to total HF rows
            if sample_remaining_rows > 0:
                hf_estimated_metaphors = (sample_metaphor_count / sample_remaining_rows) * row_count
            else:
                hf_estimated_metaphors = 0
            final_estimate_excluding_extra = hf_estimated_metaphors + lf_metaphor_count  # add LF metaphors (coded in full)

            results[sheet_name]['hf_estimated_metaphors'] = hf_estimated_metaphors
            results[sheet_name]['final_estimate_excluding_extra'] = final_estimate_excluding_extra

            
            print(f"After removing low frequency types:")
            print(f"  Sample remaining rows: {sample_remaining_rows}")
            print(f"  Sample metaphors: {sample_metaphor_count}")
    
    # --- Added: workbook-level final estimate across all sheets ---
    # Final estimate = (HF metaphors scaled up from 20% sample) + (LF metaphors counted directly) + (extra metaphors).
    total_hf_estimated = sum(
        sheet_data.get('hf_estimated_metaphors', 0)
        for sheet_name, sheet_data in results.items()
        if sheet_name != 'extra_metaphors' and isinstance(sheet_data, dict)
    )
    total_lf_metaphors = sum(
        sheet_data.get('low_freq_metaphors', 0)
        for sheet_name, sheet_data in results.items()
        if sheet_name != 'extra_metaphors' and isinstance(sheet_data, dict)
    )
    final_estimate_total = total_hf_estimated + total_lf_metaphors + results.get('extra_metaphors', 0)

    results['hf_estimated_metaphors_total'] = total_hf_estimated
    results['final_estimate_total'] = final_estimate_total


    # Save the workbook
    output_path = workbook_path.replace('.xlsx', '_processed.xlsx')
    workbook.save(output_path)
    print(f"Saved processed workbook to: {output_path}")
    
    return results

def process_directory(base_dir):
    all_results = {}
    
    # Traverse directory structure
    for root, dirs, files in os.walk(base_dir):
        for file in files:
            if file.endswith('.xlsx') and not file.endswith('_processed.xlsx'):
                filepath = os.path.join(root, file)
                try:
                    results = process_workbook(filepath)
                    all_results[filepath] = results
                except Exception as e:
                    print(f"Error processing {filepath}: {e}")
    
    return all_results

def print_tabular_results(results):
    try:
        from tabulate import tabulate
    except ImportError:
        print("Installing tabulate package...")
        import pip
        pip.main(['install', 'tabulate'])
        from tabulate import tabulate
        
    for filepath, data in results.items():
        print(f"\nFile: {os.path.basename(filepath)}")
        
        table_data = []
        # Updated headers with new columns
        headers = ["Sheet Name", "Total Rows", "Total HF Rows", "Total LF Rows", "Sample Rows (after LF removal)", "Low Freq Metaphors", "Extra Metaphors", "Sample Metaphors (after LF removal)", "HF Metaphors (scaled to 100%)", "Final Estimate (HF scaled + LF + Extra)"]
        
        # Track totals
        total_rows = 0
        total_hf_rows = 0
        total_lf_rows = 0
        total_lf_metaphors = 0
        total_sample_rows = 0
        total_sample_metaphors = 0
        total_hf_estimated_metaphors = 0
        extra_metaphors = data.get('extra_metaphors', 0)
        
        # Add rows for each sheet
        for sheet_name, sheet_data in data.items():
            # Skip 'extra_metaphors' and any coding list sheets
            if sheet_name != 'extra_metaphors' and not sheet_name.lower() == 'coding list' and not sheet_name.lower() == 'coding lists':
                # Updated row with new columns
                row = [
                    sheet_name,
                    sheet_data.get('total_rows', 0),
                    sheet_data.get('total_hf_rows', 0),
                    sheet_data.get('total_lf_rows', 0),
                    sheet_data.get('sample_remaining_rows', 0),
                    sheet_data.get('low_freq_metaphors', 0),
                    0,  # Extra metaphors (0 for individual sheets)
                    sheet_data.get('sample_metaphors', 0),
                    sheet_data.get('hf_estimated_metaphors', sheet_data.get('sample_metaphors', 0) * 5),
                    sheet_data.get('final_estimate_excluding_extra', (sheet_data.get('sample_metaphors', 0) * 5) + sheet_data.get('low_freq_metaphors', 0))
                ]
                table_data.append(row)
                
                # Add to totals
                total_rows += sheet_data.get('total_rows', 0)
                total_hf_rows += sheet_data.get('total_hf_rows', 0)
                total_lf_rows += sheet_data.get('total_lf_rows', 0)
                total_lf_metaphors += sheet_data.get('low_freq_metaphors', 0)
                total_sample_rows += sheet_data.get('sample_remaining_rows', 0)
                total_sample_metaphors += sheet_data.get('sample_metaphors', 0)
                total_hf_estimated_metaphors += sheet_data.get('hf_estimated_metaphors', sheet_data.get('sample_metaphors', 0) * 5)
        
        # Add the totals row with extra metaphors in their own column
        total_row = [
            "TOTAL", 
            total_rows,
            total_hf_rows,
            total_lf_rows,
            total_sample_rows,
            total_lf_metaphors,
            extra_metaphors,  # Extra metaphors in their own column
            total_sample_metaphors,
            total_hf_estimated_metaphors,
            (total_hf_estimated_metaphors + total_lf_metaphors + extra_metaphors)
        ]
        table_data.append(total_row)
        
        # Print the table
        print(tabulate(table_data, headers=headers, tablefmt="grid"))
        
        # Save table to CSV
        csv_path = os.path.join(os.path.dirname(filepath), f"{os.path.basename(filepath).replace('.xlsx', '')}_summary.csv")
        with open(csv_path, 'w', newline='') as csvfile:
            writer = csv.writer(csvfile)
            writer.writerow(headers)
            writer.writerows(table_data)
        print(f"Summary saved to: {csv_path}")

def save_consolidated_results(results, base_dir):
    """Save all results to a single consolidated CSV file."""
    consolidated_data = []
    # Updated headers with new columns
    headers = ["File", "Sheet Name", "Total Rows", "Total HF Rows", "Total LF Rows", "Sample Rows (after LF removal)", "Low Freq Metaphors", "Extra Metaphors", "Sample Metaphors (after LF removal)", "HF Metaphors (scaled to 100%)", "Final Estimate (HF scaled + LF + Extra)"]
    
    for filepath, data in results.items():
        filename = os.path.basename(filepath)
        extra_metaphors = data.get('extra_metaphors', 0)
        
        # Add rows for each sheet
        for sheet_name, sheet_data in data.items():
            # Skip 'extra_metaphors' and any coding list sheets
            if sheet_name != 'extra_metaphors' and not sheet_name.lower() == 'coding list' and not sheet_name.lower() == 'coding lists':
                # Updated row with new columns
                row = [
                    filename,
                    sheet_name,
                    sheet_data.get('total_rows', 0),
                    sheet_data.get('total_hf_rows', 0),
                    sheet_data.get('total_lf_rows', 0),
                    sheet_data.get('sample_remaining_rows', 0),
                    sheet_data.get('low_freq_metaphors', 0),
                    0,  # Extra metaphors (0 for individual sheets)
                    sheet_data.get('sample_metaphors', 0),
                    sheet_data.get('hf_estimated_metaphors', sheet_data.get('sample_metaphors', 0) * 5),
                    sheet_data.get('final_estimate_excluding_extra', (sheet_data.get('sample_metaphors', 0) * 5) + sheet_data.get('low_freq_metaphors', 0))
                ]
                consolidated_data.append(row)
        
        # Add the totals row with extra metaphors in their own column
        total_rows = sum(sheet_data.get('total_rows', 0) for sheet_name, sheet_data in data.items() 
                         if sheet_name != 'extra_metaphors' and not sheet_name.lower() == 'coding list' and not sheet_name.lower() == 'coding lists')
        total_hf_rows = sum(sheet_data.get('total_hf_rows', 0) for sheet_name, sheet_data in data.items() 
                            if sheet_name != 'extra_metaphors' and not sheet_name.lower() == 'coding list' and not sheet_name.lower() == 'coding lists')
        total_lf_rows = sum(sheet_data.get('total_lf_rows', 0) for sheet_name, sheet_data in data.items() 
                            if sheet_name != 'extra_metaphors' and not sheet_name.lower() == 'coding list' and not sheet_name.lower() == 'coding lists')
        total_lf_metaphors = sum(sheet_data.get('low_freq_metaphors', 0) for sheet_name, sheet_data in data.items() 
                              if sheet_name != 'extra_metaphors' and not sheet_name.lower() == 'coding list' and not sheet_name.lower() == 'coding lists')
        total_sample_rows = sum(sheet_data.get('sample_remaining_rows', 0) for sheet_name, sheet_data in data.items() 
                             if sheet_name != 'extra_metaphors' and not sheet_name.lower() == 'coding list' and not sheet_name.lower() == 'coding lists')
        total_sample_metaphors = sum(sheet_data.get('sample_metaphors', 0) for sheet_name, sheet_data in data.items() 
                                  if sheet_name != 'extra_metaphors' and not sheet_name.lower() == 'coding list' and not sheet_name.lower() == 'coding lists')
        
        # Updated total row with new columns
        total_row = [
            filename,
            "TOTAL", 
            total_rows,
            total_hf_rows,
            total_lf_rows,
            total_sample_rows,
            total_lf_metaphors,
            extra_metaphors,  # Extra metaphors in their own column
            total_sample_metaphors,
            (total_sample_metaphors * 5),
            ((total_sample_metaphors * 5) + total_lf_metaphors + extra_metaphors)
        ]
        consolidated_data.append(total_row)
        
        # Add a blank row between files for readability
        consolidated_data.append([""] * len(headers))
    
    # Save consolidated data to CSV
    consolidated_path = os.path.join(base_dir, "all_workbooks_summary.csv")
    with open(consolidated_path, 'w', newline='') as csvfile:
        writer = csv.writer(csvfile)
        writer.writerow(headers)
        writer.writerows(consolidated_data)
    print(f"\nConsolidated summary saved to: {consolidated_path}")


if __name__ == '__main__':
    # Use the same base directory as in processworkbook.py
    base_dir = 'PATH/TO/YOUR/WORKBOOKS/DIRECTORY'
    
    print(f"Processing files in: {base_dir}")
    results = process_directory(base_dir)
    
    # Print tabular results
    print_tabular_results(results)

    # Save consolidated results
    save_consolidated_results(results, base_dir)
