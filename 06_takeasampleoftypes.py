import openpyxl
import random

# Load the workbook and select the worksheet
workbook_path = 'PATH/TO/YOUR/EXCELFILE.xlsx'
print(f"Loading workbook from: {workbook_path}")
workbook = openpyxl.load_workbook(workbook_path)
worksheet_name = 'S7.1+'
print(f"Selecting worksheet: {worksheet_name}")
worksheet = workbook[worksheet_name]

# Group rows by the value in Column C (case insensitive)
rows_by_value = {}
for row in worksheet.iter_rows(min_row=2, values_only=True):
    key = str(row[2]).lower()  # Column C, case insensitive
    if key not in rows_by_value:
        rows_by_value[key] = []
    rows_by_value[key].append(row)

    # Create a new worksheet
new_worksheet_name = worksheet_name + "_20%"
print(f"Creating new worksheet: {new_worksheet_name}")
new_worksheet = workbook.create_sheet(new_worksheet_name)

    # Copy the header row
header = [cell.value for cell in worksheet[1]]
new_worksheet.append(header)

    # Sample rows and copy to the new worksheet
for key, rows in rows_by_value.items():
    if len(rows) < 20:
        sampled_rows = rows  # Keep all rows if less than 20
    else:
        sample_size = max(1, int(len(rows) * 0.2))  # Ensure at least one row is sampled
        sampled_rows = random.sample(rows, sample_size)
    for row in sampled_rows:
        new_row = []
        for cell in row:
            new_row.append(cell)
        new_worksheet.append(new_row)

# Save the workbook
workbook.save(workbook_path)
print(f"Workbook saved to: {workbook_path}")

# Print the number of rows in each worksheet
for worksheet_name in workbook.sheetnames:
    worksheet = workbook[worksheet_name]
    num_rows = worksheet.max_row
    print(f"{worksheet_name} = {num_rows} tokens")